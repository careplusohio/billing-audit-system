import psycopg2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os

# Database connection
conn = psycopg2.connect(
    dbname="careplus",
    user="postgres",
    password="careplus123",
    host="localhost",
    port="5432"
)
cur = conn.cursor()

# Fetch data
cur.execute("""
    SELECT 
        patient_first_name,
        patient_last_name,
        patient_dob,
        total_billed,
        total_paid,
        payer,
        claim_status
    FROM rhino_claims;
""")
rows = cur.fetchall()
columns = [desc[0] for desc in cur.description]

# Load into pandas DataFrame
df = pd.DataFrame(rows, columns=columns)

# Save to Excel
excel_path = "rhino_claims_selected_export.xlsx"
df.to_excel(excel_path, index=False)

# Open the file with openpyxl
wb = load_workbook(excel_path)
ws = wb.active

# Styles
header_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Format header row
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

# Format all other rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.alignment = center_alignment
        cell.border = thin_border

# Highlight unpaid claims
claim_status_col = None
for idx, cell in enumerate(ws[1], start=1):
    if cell.value == "claim_status":
        claim_status_col = idx
        break

if claim_status_col:
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        claim_status_cell = row[claim_status_col - 1]
        if claim_status_cell.value and claim_status_cell.value.lower() != "paid":
            for c in row:
                c.fill = red_fill

# Autofit column width
for column_cells in ws.columns:
    max_length = 0
    column = column_cells[0].column_letter
    for cell in column_cells:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Freeze header row
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

# Save Excel
wb.save(excel_path)
print(f"✅ Excel file saved: {excel_path}")

# Close DB connection
cur.close()
conn.close()

# Email Sending
gmail_user = 'careplusohio3@gmail.com'
gmail_password = 'kssd wdjl xaeq nrpu'  # Replace with your real App Password

to_email = 'careplusohio3@gmail.com'
subject = 'CarePlus Ohio - Rhino Claims Export 📋'
body = """
Hello,

Attached is the latest CarePlus Ohio Rhino Claims export.

Best regards,
CarePlus Ohio System
"""

msg = MIMEMultipart()
msg['From'] = gmail_user
msg['To'] = to_email
msg['Subject'] = subject
msg.attach(MIMEText(body, 'plain'))

# Attach the Excel file
if os.path.exists(excel_path):
    with open(excel_path, "rb") as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename="{os.path.basename(excel_path)}"')
    msg.attach(part)
else:
    print(f"❌ Attachment not found: {excel_path}")
    exit()

# Send the email
try:
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(gmail_user, gmail_password)
    server.sendmail(gmail_user, to_email, msg.as_string())
    server.quit()
    print("✅ Email sent successfully with the Excel file attached!")
except Exception as e:
    print(f"❌ Failed to send email: {e}")
