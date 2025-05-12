import psycopg2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

# Connect to PostgreSQL
conn = psycopg2.connect(
    dbname="careplus",
    user="postgres",
    password="careplus123",
    host="localhost",
    port="5432"
)
cur = conn.cursor()

# Fetch data
cur.execute("""
    SELECT 
        patient_first_name,
        patient_last_name,
        patient_dob,
        total_billed,
        total_paid,
        payer,
        claim_status
    FROM rhino_claims;
""")
rows = cur.fetchall()
columns = [desc[0] for desc in cur.description]

# Load into pandas DataFrame
df = pd.DataFrame(rows, columns=columns)

# Save raw data to Excel first
excel_path = "rhino_claims_selected_export.xlsx"
df.to_excel(excel_path, index=False)

# Now open with openpyxl to format
wb = load_workbook(excel_path)
ws = wb.active

# Styles
header_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Format header
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

# Format data rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.alignment = center_alignment
        cell.border = thin_border

# Autofit columns
for column_cells in ws.columns:
    max_length = 0
    column = column_cells[0].column_letter
    for cell in column_cells:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Save formatted workbook
wb.save(excel_path)

print("✅ Formatted Excel export completed: rhino_claims_selected_export.xlsx")

# Close DB connection
cur.close()
conn.close()
