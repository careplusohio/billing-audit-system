import psycopg2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os

try:
    # Connect to PostgreSQL
    conn = psycopg2.connect(
        dbname="careplus",
        user="postgres",
        password="careplus123",
        host="localhost",
        port="5432"
    )
    cur = conn.cursor()

    # Fetch all data from rhino_claims
    cur.execute("SELECT * FROM rhino_claims;")
    rows = cur.fetchall()
    columns = [desc[0] for desc in cur.description]

    # Load into pandas DataFrame
    df = pd.DataFrame(rows, columns=columns)

    # Save to Excel
    export_path = "rhino_claims_exported.xlsx"
    df.to_excel(export_path, index=False)

    # Optional: Add styles
    wb = load_workbook(export_path)
    ws = wb.active

    # Header style
    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Apply styles
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border

    # Apply border and center to all data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = thin_border

    # Autofit columns
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Add auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Save workbook
    wb.save(export_path)

    # Open the file automatically
    os.startfile(export_path)

    print("✅ rhino_claims_exported.xlsx created and opened successfully!")

except Exception as e:
    print("❌ Export failed:", e)

finally:
    if 'cur' in locals():
        cur.close()
    if 'conn' in locals():
        conn.close()
print("✅ rhino_claims_exported.xlsx created successfully!")

# Only try to open after confirmation
try:
    os.startfile(export_path)
except Exception as e:
    print("⚠️ Could not open file automatically:", e)
