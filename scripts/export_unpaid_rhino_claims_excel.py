import psycopg2
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
from datetime import datetime

# Connect to PostgreSQL
conn = psycopg2.connect(
    dbname="careplus",
    user="postgres",
    password="careplus123",
    host="localhost",
    port="5432"
)
cur = conn.cursor()

# Fetch unpaid claims only
cur.execute("""
    SELECT * FROM rhino_claims
    WHERE LOWER(claim_status) != 'paid' OR claim_status IS NULL;
""")
rows = cur.fetchall()
columns = [desc[0] for desc in cur.description]

# Load into pandas DataFrame
df = pd.DataFrame(rows, columns=columns)

# Create /exports/ folder if not exist
exports_dir = os.path.join(os.getcwd(), "exports")
os.makedirs(exports_dir, exist_ok=True)

# Add timestamp to filename
timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
excel_path = os.path.join(exports_dir, f"unpaid_rhino_claims_export_{timestamp}.xlsx")

# Save to Excel
df.to_excel(excel_path, index=False)

# Open the file with openpyxl
wb = load_workbook(excel_path)
ws = wb.active

# Styles
header_font = Font(bold=True)
center_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Format header row
for cell in ws[1]:
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

# Format all other rows
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.alignment = center_alignment
        cell.border = thin_border

# Autofit column width
for column_cells in ws.columns:
    max_length = 0
    column = column_cells[0].column_letter
    for cell in column_cells:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = (max_length + 2)
    ws.column_dimensions[column].width = adjusted_width

# Freeze the header row
ws.freeze_panes = "A2"

# Add autofilter
ws.auto_filter.ref = ws.dimensions

# Save workbook
wb.save(excel_path)

print(f"✅ Unpaid claims export completed: {excel_path}")

# Close DB connection
cur.close()
conn.close()

# Automatically open the file after saving
os.startfile(excel_path)
